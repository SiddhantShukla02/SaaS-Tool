"""
app/ui.py — Streamlit UI for the blog pipeline SaaS.

Single-file UI with two views:
  1. Dashboard      — list of runs, New Run button, metrics
  2. Run detail     — per-run progress, gate control, action buttons

Launch:
    streamlit run app/ui.py

Env vars:
    SAAS_PASSWORD          — shared login password
    SAAS_DB_PATH           — path to SQLite DB (default: runs.db)
    SAAS_QUEUE_BACKEND     — 'thread' (dev) or 'rq' (prod)
    REDIS_URL              — if backend is 'rq'
    SAAS_ALLOW_NO_AUTH=1   — bypass login for local dev
"""
import io
import os
import re
import json

import streamlit as st
from streamlit_autorefresh import st_autorefresh

# Set page config FIRST — before any other st.* call
st.set_page_config(
    page_title="SaaS for Blog — Divinheal",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="collapsed",
)

from app import orchestrator, auth          # noqa: E402
from app.repositories import state_repo as db
from app.repositories.run_repo import get_run_country_codes, get_country_codes_for_runs
from app.storage import r2_get_text
from app.repositories.search_repo import (
    get_paa_questions_for_run,
    get_serp_urls_for_run,
    save_selected_urls,
    get_selected_urls,
)
from app.repositories.output_repo import (
    get_question_bank_output,
    get_platform_draft_output,
)

# Require login
auth.require_auth()

# ─────────────────────────────────────────────────────────────
# Styling — keep it minimal, let Streamlit's defaults do the heavy lifting
# ─────────────────────────────────────────────────────────────

st.markdown("""
<style>
.status-pill { display: inline-block; padding: 3px 10px; border-radius: 10px;
               font-size: 11px; font-weight: 500; }
.pill-running { background: #E6F1FB; color: #0C447C; }
.pill-gate    { background: #FAEEDA; color: #854F0B; }
.pill-done    { background: #E1F5EE; color: #085041; }
.pill-failed  { background: #FCEBEB; color: #791F1F; }
.pill-draft   { background: #F1EFE8; color: #444441; }
.progress-bar { display: flex; gap: 4px; margin: 6px 0; }
.progress-step { flex: 1; height: 5px; border-radius: 2px; background: #D3D1C7; }
.step-done   { background: #1D9E75; }
.step-active { background: #378ADD; }
.step-failed { background: #E24B4A; }
.run-card { padding: 12px; border: 0.5px solid rgba(0,0,0,0.1); border-radius: 8px;
            margin-bottom: 10px; }
.gate-box { background: #FAEEDA; padding: 12px; border-radius: 8px;
            border-left: 4px solid #EF9F27; color: #633806; margin: 10px 0; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# Status label helpers
# ─────────────────────────────────────────────────────────────

STATUS_LABELS = {
    db.STATUS_DRAFT:              ("Draft", "pill-draft"),
    db.STATUS_STAGE1_RUNNING:     ("Stage 1 running", "pill-running"),
    db.STATUS_STAGE2_RUNNING:     ("Stage 2 running", "pill-running"),
    db.STATUS_AWAITING_FINAL_URL: ("Awaiting URL selection", "pill-gate"),
    db.STATUS_STAGE3_RUNNING:     ("Stage 3 running (Blog writing)", "pill-running"),
    db.STATUS_BLOG_READY:         ("Blog ready", "pill-done"),
    db.STATUS_BANK_RUNNING:       ("Building question bank", "pill-running"),
    db.STATUS_BANK_READY:         ("Question bank ready", "pill-done"),
    db.STATUS_QUORA_RUNNING:      ("Generating Quora", "pill-running"),
    db.STATUS_REDDIT_RUNNING:     ("Generating Reddit", "pill-running"),
    db.STATUS_SUBSTACK_RUNNING:   ("Generating Substack", "pill-running"),
    db.STATUS_COMPLETE:           ("Complete", "pill-done"),
    db.STATUS_FAILED:             ("Failed", "pill-failed"),
    db.STATUS_CANCELLED:          ("Cancelled", "pill-draft"),
}


def status_pill(status: str) -> str:
    label, cls = STATUS_LABELS.get(status, (status, "pill-draft"))
    return f'<span class="status-pill {cls}">{label}</span>'


def progress_bar_html(progress: dict) -> str:
    def cls(s):
        return {"done": "step-done", "active": "step-active",
                "failed": "step-failed"}.get(s, "")
    return (
        '<div class="progress-bar">'
        f'<div class="progress-step {cls(progress.get("stage_1", "pending"))}"></div>'
        f'<div class="progress-step {cls(progress.get("stage_2", "pending"))}"></div>'
        f'<div class="progress-step {cls(progress.get("stage_3", "pending"))}"></div>'
        f'<div class="progress-step {cls(progress.get("stage_4", "pending"))}"></div>'
        '</div>'
    )


def progress_from_run_status(run: dict) -> dict:
    status = run["status"]

    state_to_progress = {
        db.STATUS_DRAFT:              ("pending", "pending", "pending", "pending"),
        db.STATUS_STAGE1_RUNNING:     ("active",  "pending", "pending", "pending"),
        db.STATUS_STAGE2_RUNNING:     ("done",    "active",  "pending", "pending"),
        db.STATUS_AWAITING_FINAL_URL: ("done",    "done",    "pending", "pending"),
        db.STATUS_STAGE3_RUNNING:     ("done",    "done",    "active",  "pending"),
        db.STATUS_BLOG_READY:         ("done",    "done",    "done",    "pending"),
        db.STATUS_BANK_RUNNING:       ("done",    "done",    "done",    "active"),
        db.STATUS_BANK_READY:         ("done",    "done",    "done",    "done"),
        db.STATUS_QUORA_RUNNING:      ("done",    "done",    "done",    "done"),
        db.STATUS_REDDIT_RUNNING:     ("done",    "done",    "done",    "done"),
        db.STATUS_SUBSTACK_RUNNING:   ("done",    "done",    "done",    "done"),
        db.STATUS_COMPLETE:           ("done",    "done",    "done",    "done"),
        db.STATUS_CANCELLED:          ("pending", "pending", "pending", "pending"),
    }

    if status == db.STATUS_FAILED:
        failed_at_stage = run.get("failed_at_stage", "")
        marker = {
            "stage_1_serp_paa":  ("failed",  "pending", "pending", "pending"),
            "stage_2_context":   ("done",    "failed",  "pending", "pending"),
            "stage_3_blog":      ("done",    "done",    "failed",  "pending"),
            "stage_4_bank":      ("done",    "done",    "done",    "failed"),
            "stage_5_drafts":    ("done",    "done",    "done",    "done"),
        }.get(failed_at_stage, ("pending",) * 4)
    else:
        marker = state_to_progress.get(status, ("pending",) * 4)

    return {
        "stage_1": marker[0],
        "stage_2": marker[1],
        "stage_3": marker[2],
        "stage_4": marker[3],
    }


def load_dashboard_data(limit: int = 30) -> dict:
    runs = db.list_runs(limit=limit)
    run_ids = [run["id"] for run in runs]

    return {
        "metrics": orchestrator.get_dashboard_metrics(),
        "runs": runs,
        "country_codes_by_run_id": get_country_codes_for_runs(run_ids),
        "has_runs": bool(runs),
    }
# ─────────────────────────────────────────────────────────────
# Header
# ─────────────────────────────────────────────────────────────

header_left, header_right = st.columns([6, 1])
with header_left:
    st.markdown("## SaaS for Blog")
    st.caption(f"Divinheal content pipeline · signed in as **{auth.current_user()}**")
with header_right:
    st.write("")
    if st.button("Log out", use_container_width=True):
        auth.logout()

# Query param routing: ?run=42 opens detail view
run_view = st.query_params.get("run")


# ═════════════════════════════════════════════════════════════
# Run detail view
# ═════════════════════════════════════════════════════════════

def render_run_detail(run_id: int):
    run = db.get_run(run_id)
    if run is None:
        st.error(f"Run {run_id} not found")
        if st.button("← Back to dashboard"):
            st.query_params.clear()
            st.rerun()
        return

    if st.button("← Back to dashboard"):
        st.query_params.clear()
        st.rerun()

    # Title + status
    col_l, col_r = st.columns([5, 1])
    with col_l:
        st.markdown(f"### {run['primary_keyword']}")
        st.caption(
            f"Run #{run['id']} · created by {run['created_by']} · "
            f"{str(run['created_at']).replace('T', ' ')}"
        )
    with col_r:
        st.markdown(status_pill(run["status"]), unsafe_allow_html=True)

    # Info tiles
    m1, m2, m3 = st.columns(3)
    with m1:
        country_codes = get_run_country_codes(run["id"])
        st.metric("Target countries", ", ".join(country_codes) or "—")
    with m2:
        st.metric("Run ID", f"#{run['id']}")
    with m3:
        from app.database import fetch_one

        blog_row = fetch_one(
            """
            SELECT r2_key
            FROM generated_outputs
            WHERE run_id = %s AND output_type = %s
            ORDER BY id DESC
            LIMIT 1
            """,
            (run["id"], "blog"),
        )

        if blog_row:
            blog_r2_key = blog_row["r2_key"]
            st.markdown(f"**Blog:** `{blog_r2_key}`")

            try:
                final_blog_text = r2_get_text(blog_r2_key)

                st.download_button(
                    label="⬇️ Download final blog (.txt)",
                    data=final_blog_text,
                    file_name=f"run_{run['id']}_final_blog.txt",
                    mime="text/plain",
                    use_container_width=False,
                    key=f"download-final-blog-{run['id']}",
                )

            except Exception as e:
                st.warning(f"Blog found in R2, but download failed: {e}")
        else:
            st.markdown("**Blog:** *(not ready yet)*")

    # Progress
    st.markdown(
        progress_bar_html(orchestrator.progress_for_run(run_id)),
        unsafe_allow_html=True,
    )

    # Stage executions table
    st.markdown("#### Pipeline progress")
    execs = db.get_stage_executions(run_id)
    STAGE_DISPLAY = {
        "stage_1_serp_paa": "Stage 1 — Search discovery (Cells 1,3,5,7)",
        "stage_2_context":  "Stage 2 — Context build (Competitor + Forums)",
        "stage_3_blog":     "Stage 3 — Blog writing (Cells 23,25,27,29,31,33)",
        "stage_4_bank":     "Stage 4 — Question Bank",
        "stage_5_drafts":   "Stage 5 — Platform Drafts",
    }
    if not execs:
        st.info("No stages have run yet.")
    else:
        for e in execs:
            icon = {"success": "✅", "failed": "❌", "running": "🔵"}.get(
                e["status"], "⏳",
            )
            label = STAGE_DISPLAY.get(e["stage_name"], e["stage_name"])
            dur = f"{e['duration_secs']:.0f}s" if e["duration_secs"] else "—"
            st.write(f"{icon} {label} · {dur} · `{e['status']}`")
            if e.get("error_message"):
                with st.expander("Error details"):
                    st.code(e["error_message"])

    # ── Action panel — depends on run status ──
    st.markdown("#### Actions")

    if run["status"] == db.STATUS_AWAITING_FINAL_URL:
        st.markdown(
            '<div class="gate-box"><strong>Action needed:</strong> '
            'Select or paste the competitor URLs you want to use. '
            'These will be saved to the database before context extraction runs.'
            '</div>',
            unsafe_allow_html=True,
        )

        serp_rows = get_serp_urls_for_run(run_id)

        st.markdown("##### SERP URLs found")

        selected_from_serp = []

        if serp_rows:
            export_lines = []
            grouped_serp_rows = {}

            for row in serp_rows:
                group_key = (row["keyword"], row["country_code"])
                grouped_serp_rows.setdefault(group_key, []).append(row)

            for (keyword, country_code), rows in grouped_serp_rows.items():
                if export_lines:
                    export_lines.append("")

                export_lines.append(f"# {keyword} | {country_code}")

                for row in rows:
                    export_lines.append(row["url"])

            st.download_button(
                "⬇️ Download all SERP URLs as TXT",
                data="\n".join(export_lines),
                file_name=f"run_{run_id}_serp_urls.txt",
                mime="text/plain",
                use_container_width=True,
            )

            for (keyword, country_code), rows in grouped_serp_rows.items():
                st.markdown(f"**{keyword} · {country_code.upper()}**")

                for row in rows:
                    label = f"{row['rank']}. {row['url']}"
                    checked = st.checkbox(label, key=f"serp-url-{row['id']}")
                    if checked:
                        selected_from_serp.append(row["url"])

                st.divider()
        else:
            st.info("No SERP URLs found. You can still paste manual URLs below.")

        manual_urls_raw = st.text_area(
            "Additional/manual URLs, one per line",
            placeholder="https://example.com/page-1\nhttps://example.com/page-2",
            help="Paste approved URLs here after manual DA/relevance checks. One URL per line.",
            key=f"manual-urls-{run_id}",
        )

        col_a, col_b = st.columns([1, 1])
        with col_a:
            if st.button("✓ URLs ready, run Stage 2 - Context build", type="primary",
                        use_container_width=True):
                try:
                    manual_urls = re.findall(
                        r"https?://[^\s,]+",
                        manual_urls_raw,
                    )

                    manual_urls = [
                        url.strip().rstrip(".,);]")
                        for url in manual_urls
                        if url.strip()
                    ]

                    final_urls = list(dict.fromkeys(selected_from_serp + manual_urls))
                    
                    if not final_urls:
                        st.error("Please select or enter at least one URL.")
                    else:
                        save_selected_urls(run_id, final_urls)
                        orchestrator.mark_final_url_ready(run_id, auth.current_user())
                        st.success("URLs saved. Stage 2 - Context build is queued!")
                        st.rerun()

                except Exception as e:
                    st.error(str(e))
        with col_b:
            if st.button("Cancel run", use_container_width=True):
                orchestrator.cancel_run(run_id, auth.current_user())
                st.rerun()

    elif run["status"] == db.STATUS_FAILED:
        failed_stage = run.get("failed_at_stage", "unknown")

        if failed_stage == "stage_5_drafts":
            st.warning(
                "Platform draft generation failed. "
                "Your blog and question bank are still available."
            )
        else:
            st.error(f"Failed at: {failed_stage}")

        if run.get("error_message"):
            with st.expander("Error message"):
                st.code(run["error_message"])

        if st.button("🔁 Retry failed stage", type="primary"):
            try:
                orchestrator.retry_failed_stage(run_id)
                st.rerun()
            except Exception as e:
                st.error(str(e))

    elif run["status"] in (db.STATUS_BLOG_READY, db.STATUS_BANK_READY,
                             db.STATUS_COMPLETE):
        col_bank, col_q, col_r, col_s = st.columns(4)
        with col_bank:
            if run["status"] == db.STATUS_BLOG_READY:
                if st.button("📦 Build Question Bank", use_container_width=True):
                    orchestrator.start_question_bank(run_id)
                    st.rerun()
            else:
                st.write("✓ Question bank built")
        bank_ready = run["status"] in (db.STATUS_BANK_READY, db.STATUS_COMPLETE)
        with col_q:
            if st.button("✍️ Generate Quora drafts", use_container_width=True,
                          disabled=not bank_ready):
                orchestrator.start_platform_drafts(run_id, "quora")
                st.rerun()
        with col_r:
            if st.button("🟠 Generate Reddit drafts", use_container_width=True,
                          disabled=not bank_ready):
                orchestrator.start_platform_drafts(run_id, "reddit")
                st.rerun()
        with col_s:
            if st.button("🔵 Generate Substack", use_container_width=True,
                          disabled=not bank_ready):
                orchestrator.start_platform_drafts(run_id, "substack")
                st.rerun()

    elif run["status"] in (db.STATUS_STAGE1_RUNNING, db.STATUS_STAGE2_RUNNING,
                             db.STATUS_STAGE3_RUNNING, db.STATUS_BANK_RUNNING,
                             db.STATUS_QUORA_RUNNING, db.STATUS_REDDIT_RUNNING,
                             db.STATUS_SUBSTACK_RUNNING):
        st.info("🔄 Currently running. This page auto-refreshes every 2s.")

        if st.button("🛑 Stop run", type="secondary"):
            orchestrator.cancel_run(run_id, auth.current_user())
            st.warning("Stop requested. The current cell may finish, but no further cells/stages will start.")
            st.rerun()
            
        # Auto-refresh while running without blocking Streamlit rendering.
        st_autorefresh(
            interval=2000,
            key=f"run-{run_id}-autorefresh",
        )

    # ── Downloads panel ──
    st.markdown("#### Downloads")

    download_col_paa, download_col_urls, download_col_qb, = st.columns(3)

    with download_col_paa:
        paa_rows = get_paa_questions_for_run(run_id)

        if paa_rows:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "PAA Questions"

                headers = [
                    "Keyword",
                    "Country Code",
                    "Position",
                    "Question",
                    "Snippet",
                    "Source",
                    "Source URL",
                ]

                worksheet.append(headers)

                for row in paa_rows:
                    worksheet.append([
                        row.get("keyword", ""),
                        str(row.get("country_code", "") or "").upper(),
                        row.get("position", ""),
                        row.get("question", ""),
                        row.get("snippet", ""),
                        row.get("source", ""),
                        row.get("source_url", ""),
                    ])

                header_fill = PatternFill(
                    fill_type="solid",
                    fgColor="D9EAF7",
                )

                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )

                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions

                column_widths = {
                    "A": 28,  # Keyword
                    "B": 14,  # Country Code
                    "C": 10,  # Position
                    "D": 55,  # Question
                    "E": 90,  # Snippet
                    "F": 28,  # Source
                    "G": 90,  # Source URL
                }

                for column_letter, width in column_widths.items():
                    worksheet.column_dimensions[column_letter].width = width

                for row_cells in worksheet.iter_rows():
                    for cell in row_cells:
                        cell.alignment = Alignment(
                            vertical="top",
                            wrap_text=True,
                        )

                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_index].height = 60

                output = io.BytesIO()
                workbook.save(output)
                output.seek(0)

                st.download_button(
                    label="⬇️ Download PAA",
                    data=output.getvalue(),
                    file_name=f"run_{run_id}_paa_questions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"download-paa-{run_id}",
                )

            except Exception as e:
                st.button(
                    "⬇️ Download PAA",
                    disabled=True,
                    use_container_width=True,
                    key=f"download-paa-error-{run_id}",
                )
                st.warning(f"PAA found, but download failed: {e}")
        else:
            st.button(
                "⬇️ Download PAA",
                disabled=True,
                use_container_width=True,
                key=f"download-paa-disabled-{run_id}",
            )

    with download_col_urls:
        selected_urls = get_selected_urls(run_id)

        if selected_urls:
            selected_url_lines = selected_urls
            # selected_url_lines = [
            #     f"# Selected URLs for Run #{run_id}",
            #     "",
            # ]

            # for position, url in enumerate(selected_urls, start=1):
            #     selected_url_lines.append(f"{url}")

            st.download_button(
                label="⬇️ Download selected URLs",
                data="\n".join(selected_url_lines),
                file_name=f"run_{run_id}_selected_urls.txt",
                mime="text/plain",
                use_container_width=True,
                key=f"download-selected-urls-{run_id}",
            )
        else:
            st.button(
                "⬇️ Download selected URLs",
                disabled=True,
                use_container_width=True,
                key=f"download-selected-urls-disabled-{run_id}",
            )

    with download_col_qb:
        question_bank_row = get_question_bank_output(run_id)

        if question_bank_row:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                question_bank_raw = r2_get_text(question_bank_row["r2_key"])
                question_bank_payload = json.loads(question_bank_raw or "{}")

                headers = question_bank_payload.get("headers", [])
                rows = question_bank_payload.get("rows", [])

                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Question Bank"

                if headers:
                    worksheet.append(headers)

                for row_values in rows:
                    worksheet.append([
                        "" if value is None else value
                        for value in row_values
                    ])

                header_fill = PatternFill(
                    fill_type="solid",
                    fgColor="D9EAF7",
                )

                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )

                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions

                text_heavy_columns = {
                    "Question",
                    "Original_Fragment",
                    "Competitor_Answer_Ref",
                }

                for column_cells in worksheet.columns:
                    column_letter = get_column_letter(column_cells[0].column)
                    header_value = str(column_cells[0].value or "")

                    if header_value in text_heavy_columns:
                        worksheet.column_dimensions[column_letter].width = 45
                    else:
                        max_length = max(
                            len(str(cell.value or ""))
                            for cell in column_cells[:100]
                        )
                        worksheet.column_dimensions[column_letter].width = min(
                            max(max_length + 2, 12),
                            28,
                        )

                    for cell in column_cells:
                        cell.alignment = Alignment(
                            vertical="top",
                            wrap_text=True,
                        )

                for row in worksheet.iter_rows(min_row=2):
                    worksheet.row_dimensions[row[0].row].height = 45

                output = io.BytesIO()
                workbook.save(output)
                output.seek(0)

                st.download_button(
                    label="⬇️ Download Question Bank",
                    data=output.getvalue(),
                    file_name=f"run_{run_id}_question_bank.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"download-question-bank-{run_id}",
                )

            except Exception as e:
                st.button(
                    "⬇️ Download Question Bank",
                    disabled=True,
                    use_container_width=True,
                    key=f"download-question-bank-error-{run_id}",
                )
                st.warning(f"Question bank found, but download failed: {e}")
        else:
            st.button(
                "⬇️ Download Question Bank",
                disabled=True,
                use_container_width=True,
                key=f"download-question-bank-disabled-{run_id}",
            )

    def build_platform_xlsx(output_type: str, r2_key: str) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        raw_text = r2_get_text(r2_key)
        drafts = json.loads(raw_text or "[]")

        if isinstance(drafts, dict):
            drafts = [drafts]

        platform_sheet_names = {
            "quora_drafts": "Quora Drafts",
            "reddit_drafts": "Reddit Drafts",
            "substack_drafts": "Substack Drafts",
        }

        preferred_headers = {
            "quora_drafts": [
                "Question",
                "Target_Country",
                "Funnel_Stage",
                "Priority",
                "Word_Count",
                "Draft_Markdown",
            ],
            "reddit_drafts": [
                "Question",
                "Suggested_Subreddit",
                "Target_Country",
                "Funnel_Stage",
                "Priority",
                "Word_Count",
                "Draft_Markdown",
            ],
            "substack_drafts": [
                "Headline",
                "Theme",
                "Target_Country",
                "Funnel_Stage",
                "Questions_Covered",
                "Word_Count",
                "Draft_Markdown",
            ],
        }

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = platform_sheet_names.get(output_type, "Platform Drafts")

        base_headers = preferred_headers.get(output_type, [])

        extra_headers = []
        for draft in drafts:
            if isinstance(draft, dict):
                for key in draft.keys():
                    if key not in base_headers and key not in extra_headers:
                        extra_headers.append(key)

        headers = base_headers + extra_headers

        if not headers:
            headers = ["Draft"]

        worksheet.append(headers)

        for draft in drafts:
            if isinstance(draft, dict):
                worksheet.append([
                    "" if draft.get(header) is None else draft.get(header)
                    for header in headers
                ])
            else:
                worksheet.append([str(draft)])

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        text_heavy_columns = {
            "Question",
            "Draft_Markdown",
            "Questions_Covered",
            "Suggested_Subreddit",
            "Theme",
            "Headline",
        }

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            header_value = str(column_cells[0].value or "")

            if header_value == "Draft_Markdown":
                worksheet.column_dimensions[column_letter].width = 90
            elif header_value in text_heavy_columns:
                worksheet.column_dimensions[column_letter].width = 45
            else:
                max_length = max(
                    len(str(cell.value or ""))
                    for cell in column_cells[:100]
                )
                worksheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 12),
                    28,
                )

            for cell in column_cells:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for row in worksheet.iter_rows(min_row=2):
            worksheet.row_dimensions[row[0].row].height = 80

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output.getvalue()

    def render_platform_download(
        label: str,
        platform: str,
        output_type: str,
        file_suffix: str,
        key_suffix: str,
    ) -> None:
        output_row = get_platform_draft_output(run_id, platform)

        if output_row:
            try:
                export_xlsx = build_platform_xlsx(output_type, output_row["r2_key"])

                st.download_button(
                    label=label,
                    data=export_xlsx,
                    file_name=f"run_{run_id}_{file_suffix}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"download-{key_suffix}-{run_id}",
                )
            except Exception as e:
                st.button(
                    label,
                    disabled=True,
                    use_container_width=True,
                    key=f"download-{key_suffix}-error-{run_id}",
                )
                st.warning(f"{label.replace('⬇️ ', '')} found, but download failed: {e}")
        else:
            st.button(
                label,
                disabled=True,
                use_container_width=True,
                key=f"download-{key_suffix}-disabled-{run_id}",
            )

    st.write("")

    platform_col_quora, platform_col_reddit, platform_col_substack = st.columns(3)

    with platform_col_quora:
        render_platform_download(
            label="⬇️ Download Quora",
            platform="quora",
            output_type="quora_drafts",
            file_suffix="quora_drafts",
            key_suffix="quora",
        )

    with platform_col_reddit:
        render_platform_download(
            label="⬇️ Download Reddit",
            platform="reddit",
            output_type="reddit_drafts",
            file_suffix="reddit_drafts",
            key_suffix="reddit",
        )

    with platform_col_substack:
        render_platform_download(
            label="⬇️ Download Substack",
            platform="substack",
            output_type="substack_drafts",
            file_suffix="substack_drafts",
            key_suffix="substack",
        )

    # ── Currently running stage ──
    running_exec = next((e for e in execs if e["status"] == "running"), None)

    if running_exec:
        stage_label = STAGE_DISPLAY.get(
            running_exec["stage_name"],
            running_exec["stage_name"]
        )

        st.markdown(f"### 🔄 Currently running: **{stage_label}**")
    # ── Activity log ──
    st.markdown("---")
    st.markdown("#### Activity")
    activity = db.get_activity(run_id, limit=150)
    if not activity:
        st.caption("No activity yet.")
    else:
        for a in (activity):
            prefix = {"error": "🔴", "warn": "🟡", "info": "·"}.get(a["level"], "·")
            from datetime import datetime
            from zoneinfo import ZoneInfo

            timestamp_text = str(a["timestamp"])

            try:
                dt = datetime.fromisoformat(timestamp_text.replace("Z", "+00:00"))
                dt_ist = dt.astimezone(ZoneInfo("Asia/Kolkata"))
                ts = dt_ist.strftime("%H:%M:%S")
            except Exception:
                ts = timestamp_text[11:19]  # fallback

            msg = a["message"]

            formatted = f"[{ts}] {msg}"

            if "▶ started" in msg:
                st.info(formatted)
            elif "✅ finished" in msg:
                st.success(formatted)
            elif "❌" in msg or "failed" in msg.lower() or a["level"] == "error":
                st.error(formatted)
            else:
                st.markdown(
                    f"<code>[{ts}]</code> {prefix} {msg}",
                    unsafe_allow_html=True,
                )


# ═════════════════════════════════════════════════════════════
# Dashboard view
# ═════════════════════════════════════════════════════════════

def render_dashboard():
    metrics = orchestrator.get_dashboard_metrics()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Active runs", metrics["active"])
    m2.metric("Awaiting gate", metrics["awaiting"])
    m3.metric("Completed", metrics["complete"])
    m4.metric("Cost today", f"${metrics['cost_today']:.2f}")

    st.divider()

       # New run form
    with st.expander("➕ Start new run", expanded=not bool(db.list_runs(limit=1))):
        st.markdown(
            "**Step 1:** Add one keyword-country pair per row.  \n"
            "**Step 2:** Click Start to begin Stage 1 - Search discovery."
        )
        st.caption(
            "Keyword text is preserved exactly except leading/trailing spaces. "
            "Country codes are normalized to lowercase."
        )

        default_keyword_rows = [
            {"keyword": "", "country_code": ""},
            {"keyword": "", "country_code": ""},
            {"keyword": "", "country_code": ""},
        ]

        with st.form("new_run"):
            keyword_table = st.data_editor(
                default_keyword_rows,
                column_config={
                    "keyword": st.column_config.TextColumn(
                        "Keyword",
                        help="Exact search keyword/topic to run",
                        required=False,
                    ),
                    "country_code": st.column_config.TextColumn(
                        "Country code",
                        help="2-letter country code, e.g. uk, ae, ng",
                        max_chars=2,
                        required=False,
                    ),
                },
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                key="keyword_country_editor",
            )

            submit = st.form_submit_button("Start run", type="primary")

        if submit:
            keyword_rows = []

            for row in keyword_table:
                keyword = str(row.get("keyword", "")).strip()
                country_code = str(row.get("country_code", "")).strip().lower()

                if not keyword and not country_code:
                    continue

                keyword_rows.append({
                    "keyword": keyword,
                    "country_code": country_code,
                })

            if not keyword_rows:
                st.error("Add at least one keyword-country pair.")
            else:
                try:
                    run_id = orchestrator.start_run(
                        keyword_rows,
                        auth.current_user(),
                    )
                    st.success(f"Run #{run_id} queued — Stage 1 starting")
                    st.query_params["run"] = str(run_id)
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

    # Runs list
    st.markdown("#### Runs")
    runs = db.list_runs(limit=30)
    if not runs:
        st.info("No runs yet. Click 'Start new run' above.")
        return

    run_ids = [run["id"] for run in runs]
    country_codes_by_run_id = get_country_codes_for_runs(run_ids)

    for run in runs:
        with st.container():
            cols = st.columns([4, 1.5, 1])
            with cols[0]:
                st.markdown(f"**{run['primary_keyword']}**")
                country_codes = country_codes_by_run_id.get(run["id"], [])

                st.caption(
                    f"Run #{run['id']} · {', '.join(country_codes) or '—'} · "
                    f"{str(run['created_at']).replace('T', ' ')[:16]}"
                )
                st.markdown(
                    progress_bar_html(progress_from_run_status(run)),
                    unsafe_allow_html=True,
                )
            with cols[1]:
                st.markdown(status_pill(run["status"]), unsafe_allow_html=True)
                st.caption("")
            with cols[2]:
                if st.button("Open", key=f"open-{run['id']}", use_container_width=True):
                    st.query_params["run"] = str(run["id"])
                    st.rerun()
            st.markdown("---")

# ─────────────────────────────────────────────────────────────
# Route
# ─────────────────────────────────────────────────────────────

if run_view:
    try:
        run_id_int = int(run_view)
    except ValueError:
        st.error(f"Invalid run id: {run_view}")
        if st.button("← Back to dashboard"):
            st.query_params.clear()
            st.rerun()
    else:
        render_run_detail(run_id_int)
else:
    render_dashboard()